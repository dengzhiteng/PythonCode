# 参考文档 https://www.cnblogs.com/JingYanC/p/15691853.html
from openpyxl import load_workbook
from openpyxl import Workbook
import time

filePath= "openpyxl/openpyxl-create.xlsx"

# 创建表格
# wb = Workbook()
# ws = wb.active
# ws.title = "10月工资表"
# ws.sheet_properties.tabColor = "ff0033" #红色
# wb.save(filePath)

# 以只读的形式打开
# wb = load_workbook(filePath,read_only=True)
wb = load_workbook(filePath)
ws10 = wb['10月工资表']

# 读取表格文件
# sheet_names=wb.sheetnames
# for rows in ws10['A1':'C2']:
#    for cell in rows:
#       # 修改值
#       cell.value="???"
#       print(cell.value)
# wb.save(filePath)

# 数据追加
# data = [['zhangsan@163.com','张三','男','酱油一部','打酱油',100],
#         ['wangwu@163.com','王五','男','酱油一部','打酱油',100],
#       ]
# for item in data:
#     ws10.append(item)
# wb.save(filePath)


max_row = ws10.max_row
max_col = ws10.max_column
# 获取ws10表中2行的值
# for x in range(1,max_col):
#    cell = ws10.cell(row=2, column=x).value
#    print(cell)

# 遍历ws10中单元格的值
# for row in range(2,max_row+1):
#    for col in range(1, max_col+1):
#       cell = ws10.cell(row=row, column=col).value
#       print(cell)

# 使用切片遍历单元格
# for row in ws10['A2':'F2']:
#     for cell in row:
#         print(cell.value)

#
# 单独指定行、列、或者行列的遍历范围  遍历行
# for row in ws10.iter_rows(min_row=2,max_row=3,min_col=1,max_col=3,):
#     for cell in row:
#         print(cell.value)

# 单独指定行、列、或者行列的遍历范围  遍历列
# for col in ws10.iter_cols(min_row=1, max_col=3, max_row=2):
#     for cell in col:
#         print(cell.value)

# 遍历所有行
# for row in ws10.rows:
#     for cell in row:
#         print(cell.value)

# 遍历所有列
# for col in ws10.columns:
#     for cell in col:
#         print(cell.value)

# 插入一行
# ws10.insert_rows(1)
# ws10.merge_cells('A1:F1')
# ws10['A1']='10月工资明细'


# 删除一行 删除合并的行,有点问题
ws10.delete_rows(1)
wb.save(filePath)
















