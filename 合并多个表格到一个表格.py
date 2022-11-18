from openpyxl import load_workbook,Workbook
import oslib

listDir = oslib.listdir('./openpyxl/学科')
allData = []
new_wb=Workbook()
new_ws = new_wb.active
new_ws.title = "汇总"

# 汇总到一张表格上
for fileName in listDir:
    wb = load_workbook(f'./openpyxl/学科/{fileName}')
    ws = wb.active
    for row in ws.rows:
        row_data = []
        for cell in row:
            row_data.append(cell.value)
        if row_data not in allData:
            allData.append(row_data)

for data in allData:
    new_ws.append(data)

new_wb.save('./openpyxl/成绩汇总.xlsx')