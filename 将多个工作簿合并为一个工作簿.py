from openpyxl import load_workbook,Workbook
import os

filePath='./openpyxl/成绩汇总.xlsx'
wb = load_workbook(filePath)

listDir = os.listdir('./openpyxl/学科')
for fileName in listDir:
    wb1 = load_workbook(f'./openpyxl/学科/{fileName}')
    ws1 = wb1.active
    ws=wb.create_sheet(fileName[:-5])
    for row in ws1.rows:
        row_data = []
        for cell in row:
            row_data.append(cell.value)
        ws.append(row_data)
wb.save(filePath)


# 12312


