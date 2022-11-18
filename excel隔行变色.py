from openpyxl import load_workbook
from openpyxl.styles import PatternFill, colors

filePath='./openpyxl/成绩汇总.xlsx'
wb = load_workbook(filePath)
ws = wb['汇总']
max_row = ws.max_row
max_col = ws.max_column
print(max_col,)
fill_1 = PatternFill("solid", fgColor="1874CD")

for row in range(2,max_row+1,2):
   for col in range(1, max_col + 1,):
      cell = ws.cell(row=row, column=col)
      cell.fill = fill_1

wb.save(filePath)

