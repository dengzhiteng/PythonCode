from openpyxl import load_workbook
from openpyxl import Workbook

filePath='./openpyxl/成绩单.xlsx'
wb = load_workbook(filePath)

# 语文 chinese
# 英语 english
# 数学 math
chinese_sheet= wb['语文']
english_sheet= wb['英语']
math_sheet= wb['数学']
class_sheet = wb['班级名单']
subs_sheets=[chinese_sheet,english_sheet,math_sheet]

# 获取指定科目成绩
def getScore(sheet,name):
    max_row = sheet.max_row
    max_col = sheet.max_column
    score=0
    for row in range(2,max_row+1):
       for col in range(1, max_col+1):
          cell = sheet.cell(row=row, column=col).value
          if cell==name:
            score=sheet.cell(row=row,column=col+1).value
    return score

# 获取科目对应的姓名
def getNameList(sheet):
    list=[]
    for row in sheet.iter_cols(min_row=2,min_col=1, max_col=1):
        for cell in row:
            list.append(cell.value)
    return list

# 创建 平均分表
average_sheet = wb.create_sheet("平均分")
average_sheet.sheet_properties.tabColor = 'ff72BA'
nameList=getNameList(class_sheet)

scoreList=[['姓名','总分','平均分']]
print('*'*10,'计算中......')
for name in nameList:
    total=0
    for sheet in subs_sheets:
        total +=getScore(sheet,name)

    scoreList.append([name,total,'%.2f'% (total/3)])
 # 写入表格中
for data in scoreList:
    average_sheet.append(data)

print('*'*10,'完成计算......')
wb.save(filePath)
wb.close()
